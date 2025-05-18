import openpyxl
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from src.database.connection import get_connection

def cars_data():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            cars.vin AS "VIN",
            brands.name AS "Марка",
            models.name AS "Модель",
            colors.name AS "Цвет",
            transmissions.name AS "Тип трансмиссии",
            cars.year AS "Год выпуска",
            cars.mileage AS "Пробег",
            cars.price AS "Цена",
            status.name AS "Статус"
        FROM cars
        JOIN brands ON cars.brand_id = brands.id
        JOIN models ON cars.model_id = models.id
        JOIN colors ON cars.color_id = colors.id
        JOIN transmissions ON cars.transmission_id = transmissions.id
        JOIN status ON cars.status_id = status.id;
    """)
    res = cur.fetchall()
    col = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    return res, col

def clients_data():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            full_name AS "ФИО",
            phone AS "Телефон",
            email AS "Почта"
        FROM clients
    """)
    res = cur.fetchall()
    col = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    return res, col

def sales_data():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(""" 
        SELECT 
            clients.full_name AS "Клиент",
            cars.vin AS "VIN",
            CONCAT(brands.name, ' ', models.name, ' ', colors.name) AS "Автомобиль",
            sales.date AS "Дата продажи",
            sales.price AS "Цена"
        FROM sales
        JOIN cars ON sales.car_id = cars.id
        JOIN clients ON sales.client_id = clients.id
        JOIN brands ON cars.brand_id = brands.id
        JOIN models ON cars.model_id = models.id
        JOIN colors ON cars.color_id = colors.id
    """)

    res = cur.fetchall()
    col = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    return res, col

def users_data():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            login AS "Логин",
            password AS "Пароль",
            role AS "Роль"
        FROM users;
    """)
    res = cur.fetchall()
    col = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    return res, col

def delete_car_by_vin(vin):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM cars WHERE vin = %s", (vin,))
    conn.commit()

    cur.close()
    conn.close()

def get_car_by_vin(vin):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT cars.vin, brands.name, models.name, colors.name,
               transmissions.name, cars.year, cars.mileage, cars.price, status.name
        FROM cars
        JOIN brands ON cars.brand_id = brands.id
        JOIN models ON cars.model_id = models.id
        JOIN colors ON cars.color_id = colors.id
        JOIN transmissions ON cars.transmission_id = transmissions.id
        JOIN status ON cars.status_id = status.id
        WHERE cars.vin = %s
    """, (vin,))

    car_data = cur.fetchone()
    conn.close()

    if car_data:
        return car_data
    else:
        raise ValueError("Автомобиль с таким VIN не найден")

def update_car(vin, brand, model, color, transmission, year, mileage, price):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT id FROM brands WHERE name = %s", (brand,))
    brand_id = cur.fetchone()[0]

    cur.execute("SELECT id FROM models WHERE name = %s", (model,))
    model_id = cur.fetchone()[0]

    cur.execute("SELECT id FROM colors WHERE name = %s", (color,))
    color_id = cur.fetchone()[0]

    cur.execute("SELECT id FROM transmissions WHERE name = %s", (transmission,))
    transmission_id = cur.fetchone()[0]

    cur.execute("""
        UPDATE cars
        SET brand_id = %s, model_id = %s, color_id = %s, transmission_id = %s,
            year = %s, mileage = %s, price = %s
        WHERE vin = %s
    """, (brand_id, model_id, color_id, transmission_id, year, mileage, price, vin))

    conn.commit()
    cur.close()
    conn.close()

def get_all_filters():
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT name FROM brands")
    brands = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT name FROM models")
    models = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT name FROM colors")
    colors = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT name FROM transmissions")
    transmissions = [row[0] for row in cur.fetchall()]

    cur.execute("SELECT name FROM status")
    status = [row[0] for row in cur.fetchall()]

    conn.close()

    return {
        "brands": brands,
        "models": models,
        "colors": colors,
        "transmissions": transmissions,
        "status": status,
    }

def filter_cars(filters: dict):
    conn = get_connection()
    cur = conn.cursor()

    query = """
    SELECT cars.vin, brands.name, models.name, colors.name,
           transmissions.name, cars.year, cars.mileage, cars.price, status.name
    FROM cars
    JOIN brands ON cars.brand_id = brands.id
    JOIN models ON cars.model_id = models.id
    JOIN colors ON cars.color_id = colors.id
    JOIN transmissions ON cars.transmission_id = transmissions.id
    JOIN status ON cars.status_id = status.id
    WHERE 1=1
    """
    params = []

    if filters["brand"]:
        query += " AND brands.name = %s"
        params.append(filters["brand"])
    if filters["model"]:
        query += " AND models.name = %s"
        params.append(filters["model"])
    if filters["color"]:
        query += " AND colors.name = %s"
        params.append(filters["color"])
    if filters["transmission"]:
        query += " AND transmissions.name = %s"
        params.append(filters["transmission"])
    if filters["status"]:
        query += " AND status.name = %s"
        params.append(filters["status"])
    if filters["mileage_min"] is not None:
        query += " AND cars.mileage >= %s"
        params.append(filters["mileage_min"])
    if filters["mileage_max"] is not None:
        query += " AND cars.mileage <= %s"
        params.append(filters["mileage_max"])
    if filters["year_min"] is not None:
        query += " AND cars.year >= %s"
        params.append(filters["year_min"])
    if filters["year_max"] is not None:
        query += " AND cars.year <= %s"
        params.append(filters["year_max"])

    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    conn.close()

    col_names = [
        "VIN", "Марка", "Модель", "Цвет", "Тип трансмиссии",
        "Год выпуска", "Пробег", "Цена", "Статус"
    ]
    return rows, col_names

def get_client_id(full_name, phone, email):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT id FROM clients  WHERE full_name = %s AND phone = %s AND email = %s", (full_name, phone, email))
    result = cur.fetchone()

    cur.close()
    conn.close()

    if result:
        return result[0]
    else:
        raise ValueError("Клиент не найден")

def delete_client_by_full_name(full_name):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM clients WHERE full_name = %s", (full_name,))
    conn.commit()

    cur.close()
    conn.close()

def filter_clients_universal(query: str):
    conn = get_connection()
    cur = conn.cursor()

    sql = """
        SELECT full_name AS "ФИО", phone AS "Телефон", email AS "Почта"
        FROM clients
        WHERE full_name ILIKE %s
           OR phone ILIKE %s
           OR email ILIKE %s
    """
    pattern = f"%{query}%"
    cur.execute(sql, (pattern, pattern, pattern))

    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]

    cur.close()
    conn.close()

    return rows, columns

def filter_sales(filters: dict):
    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT 
            clients.full_name,
            cars.vin,
            CONCAT(brands.name, ' ', models.name, ' ', colors.name) AS car_info,
            sales.date,
            sales.price
        FROM sales
        JOIN clients ON sales.client_id = clients.id
        JOIN cars ON sales.car_id = cars.id
        JOIN brands ON cars.brand_id = brands.id
        JOIN models ON cars.model_id = models.id
        JOIN colors ON cars.color_id = colors.id
        WHERE 1=1
    """
    params = []

    if filters["client"]:
        query += " AND clients.full_name = %s"
        params.append(filters["client"])
    if filters["date_from"]:
        query += " AND sales.date >= %s"
        params.append(filters["date_from"])
    if filters["date_to"]:
        query += " AND sales.date <= %s"
        params.append(filters["date_to"])
    if filters["price_from"]:
        query += " AND sales.price >= %s"
        params.append(filters["price_from"])
    if filters["price_to"]:
        query += " AND sales.price <= %s"
        params.append(filters["price_to"])

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    col_names = ["Клиент", "VIN", "Автомобиль", "Дата продажи", "Цена"]
    return rows, col_names

def generate_sales_report(filters=None):
    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT 
            clients.full_name,
            cars.vin,
            CONCAT(brands.name, ' ', models.name, ' ', colors.name) AS car_info,
            sales.date,
            sales.price
        FROM sales
        JOIN clients ON sales.client_id = clients.id
        JOIN cars ON sales.car_id = cars.id
        JOIN brands ON cars.brand_id = brands.id
        JOIN models ON cars.model_id = models.id
        JOIN colors ON cars.color_id = colors.id
        WHERE 1=1
    """
    params = []

    if filters:
        if filters.get("client"):
            query += " AND clients.full_name = %s"
            params.append(filters["client"])
        if filters.get("date_from"):
            query += " AND sales.date >= %s"
            params.append(filters["date_from"])
        if filters.get("date_to"):
            query += " AND sales.date <= %s"
            params.append(filters["date_to"])
        if filters.get("price_from"):
            query += " AND sales.price >= %s"
            params.append(filters["price_from"])
        if filters.get("price_to"):
            query += " AND sales.price <= %s"
            params.append(filters["price_to"])

    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    col_names = ["Клиент", "VIN", "Автомобиль", "Дата продажи", "Цена"]

    file_path, _ = QFileDialog.getSaveFileName(None, "Сохранить отчет", "", "Excel Files (*.xlsx)")

    if not file_path:
        QMessageBox.warning(None, "Ошибка", "Не выбран путь для сохранения отчета.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Продажи"

    for col_num, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = col_name

    for row_num, row_data in enumerate(rows, start=2):
        for col_num, value in enumerate(row_data, start=1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = value

    wb.save(file_path)

    conn.close()
    QMessageBox.information(None, "Отчет", f"Отчет успешно сгенерирован и сохранен по пути: {file_path}")

def get_sales_stats(filters: dict):
    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT sales.date, SUM(sales.price)
        FROM sales
        JOIN clients ON sales.client_id = clients.id
        JOIN cars ON sales.car_id = cars.id
        WHERE 1=1
    """
    params = []

    if filters.get("client"):
        query += " AND clients.full_name = %s"
        params.append(filters["client"])
    if filters.get("date_from"):
        query += " AND sales.date >= %s"
        params.append(filters["date_from"])
    if filters.get("date_to"):
        query += " AND sales.date <= %s"
        params.append(filters["date_to"])
    if filters.get("price_from"):
        query += " AND sales.price >= %s"
        params.append(filters["price_from"])
    if filters.get("price_to"):
        query += " AND sales.price <= %s"
        params.append(filters["price_to"])

    query += " GROUP BY sales.date ORDER BY sales.date"

    cur.execute(query, tuple(params))
    result = cur.fetchall()

    cur.close()
    conn.close()
    return result